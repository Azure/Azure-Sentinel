#!/usr/bin/env python3
"""Build connector_history.xlsx from connector_history.csv.

Produces a workbook with:
* "Data"  — the full CSV, styled header, frozen panes, auto-filter.
* "Stock" — a line chart of active / deprecated / total connectors over time.
* "Flow"  — a column chart of connectors created vs updated per month.

Charts reference the Data sheet live, so editing the data updates the charts.
"""
from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
CSV_PATH = HERE / "connector_history.csv"
XLSX_PATH = HERE / "connector_history.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def main() -> int:
    with open(CSV_PATH, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)
    header, data = rows[0], rows[1:]

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Header.
    ws.append(header)
    for col_idx, _ in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Data, coercing numeric columns to int.
    numeric_cols = {
        header.index(c)
        for c in (
            "active_connectors", "deprecated_connectors", "total_connectors",
            "connectors_created", "connectors_updated",
        )
        if c in header
    }
    for record in data:
        out = []
        for i, value in enumerate(record):
            if i in numeric_cols and value != "":
                out.append(int(value))
            else:
                out.append(value)
        ws.append(out)

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(header))
    ws.auto_filter.ref = f"A1:{last_col}{len(data) + 1}"
    for col_idx, name in enumerate(header, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(name) + 2)

    n_rows = len(data)
    cats = Reference(ws, min_col=header.index("month") + 1,
                     min_row=2, max_row=n_rows + 1)

    # Stock chart (line).
    stock_ws = wb.create_sheet("Stock")
    stock = LineChart()
    stock.title = "Connectors over time (as of 1st of month)"
    stock.y_axis.title = "Connectors"
    stock.x_axis.title = "Month"
    stock.height = 12
    stock.width = 28
    # openpyxl defaults axes to delete=True, which hides tick labels/titles.
    stock.x_axis.delete = False
    stock.y_axis.delete = False
    stock.legend.position = "b"
    for name in ("active_connectors", "deprecated_connectors", "total_connectors"):
        col = header.index(name) + 1
        ref = Reference(ws, min_col=col, min_row=1, max_row=n_rows + 1)
        stock.add_data(ref, titles_from_data=True)
    stock.set_categories(cats)
    stock_ws.add_chart(stock, "B2")

    # Flow chart (column) — only if the flow columns exist.
    if "connectors_created" in header and "connectors_updated" in header:
        flow_ws = wb.create_sheet("Flow")
        flow = BarChart()
        flow.type = "col"
        flow.grouping = "clustered"
        flow.title = "Connectors created vs updated per month (merges to master)"
        flow.y_axis.title = "Distinct connectors"
        flow.x_axis.title = "Month"
        flow.height = 12
        flow.width = 28
        flow.x_axis.delete = False
        flow.y_axis.delete = False
        flow.legend.position = "b"
        for name in ("connectors_created", "connectors_updated"):
            col = header.index(name) + 1
            ref = Reference(ws, min_col=col, min_row=1, max_row=n_rows + 1)
            flow.add_data(ref, titles_from_data=True)
        flow.set_categories(cats)
        flow_ws.add_chart(flow, "B2")

    wb.save(XLSX_PATH)
    print(f"Wrote {XLSX_PATH}")
    print(f"Sheets: {wb.sheetnames}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
